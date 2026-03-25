#!/usr/bin/env python3
"""
見積もり Excel ファイル生成スクリプト

使い方:
    python generate_excel.py --input <json_path> --output <xlsx_path>

入力 JSON 形式:
    {
        "project_name": "プロジェクト名",
        "client_name": "クライアント名",
        "created_date": "YYYY-MM-DD",
        "valid_until": "YYYY-MM-DD",
        "items": [
            {
                "phase": "設計|バックエンド開発|テスト / QA|プロジェクト管理",
                "task": "作業内容",
                "man_days": 2.0,
                "remarks": ""
            }
        ]
    }
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl が見つかりません。`pip install openpyxl` を実行してください。", file=sys.stderr)
    sys.exit(1)


# ─── 定数 ────────────────────────────────────────────────────────────────────

UNIT_PRICES = {
    "設計": 80_000,
    "バックエンド開発": 80_000,
    "テスト / QA": 60_000,
    "テスト/QA": 60_000,
    "プロジェクト管理": 80_000,
}

PHASE_COLORS = {
    "設計": "D6E4F0",
    "バックエンド開発": "D5F5E3",
    "テスト / QA": "FEF9E7",
    "テスト/QA": "FEF9E7",
    "プロジェクト管理": "F9EBEA",
}

PHASE_ORDER = ["設計", "バックエンド開発", "テスト / QA", "テスト/QA", "プロジェクト管理"]

HEADER_COLOR = "2E4057"
SUB_HEADER_COLOR = "607D8B"
TOTAL_COLOR = "E8F5E9"


# ─── ヘルパー関数 ───────────────────────────────────────────────────────────────

def make_thin_border():
    thin = Side(style="thin", color="999999")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def make_medium_border():
    medium = Side(style="medium", color="555555")
    thin = Side(style="thin", color="999999")
    return Border(left=medium, right=medium, top=thin, bottom=thin)


def make_header_border():
    medium = Side(style="medium", color="FFFFFF")
    return Border(left=medium, right=medium, top=medium, bottom=medium)


def apply_cell(ws, row, col, value, *, bold=False, font_size=11,
               fill_color=None, number_format=None, align="left",
               font_color="000000", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Meiryo UI", size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    if fill_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
    if number_format:
        cell.number_format = number_format
    cell.border = make_thin_border()
    return cell


def normalize_phase(phase: str) -> str:
    """フェーズ名を正規化 (テスト/QA の表記ゆれ吸収)"""
    if "テスト" in phase:
        return "テスト / QA"
    return phase


# ─── メイン関数 ─────────────────────────────────────────────────────────────────

def generate_excel(data: dict, output_path: str):
    project_name = data.get("project_name", "プロジェクト名未設定")
    client_name = data.get("client_name", "")
    created_date = data.get("created_date", datetime.today().strftime("%Y-%m-%d"))
    valid_until = data.get("valid_until", "")
    if not valid_until:
        valid_dt = datetime.strptime(created_date, "%Y-%m-%d") + timedelta(days=30)
        valid_until = valid_dt.strftime("%Y-%m-%d")

    items = data.get("items", [])

    # フェーズ名を正規化
    for item in items:
        item["phase"] = normalize_phase(item["phase"])

    # 工程の順序に従ってソート
    def phase_sort_key(item):
        normalized_order = [normalize_phase(p) for p in PHASE_ORDER]
        try:
            return normalized_order.index(item["phase"])
        except ValueError:
            return len(PHASE_ORDER)

    items_sorted = sorted(items, key=phase_sort_key)

    # ─── ワークブック作成 ────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "見積もり書"

    # 列幅設定
    col_widths = [4, 20, 40, 12, 14, 14, 24]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 行の高さ
    ws.row_dimensions[1].height = 10  # 余白

    current_row = 2

    # ─── タイトル ────────────────────────
    ws.merge_cells(f"B{current_row}:G{current_row}")
    title_cell = ws.cell(row=current_row, column=2, value="ソフトウェア開発費用見積もり書")
    title_cell.font = Font(name="Meiryo UI", size=18, bold=True, color="2E4057")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[current_row].height = 36
    current_row += 1

    ws.row_dimensions[current_row].height = 6
    current_row += 1

    # ─── プロジェクト情報 ─────────────────
    info_rows = [
        ("プロジェクト名", project_name),
        ("お客様名", client_name),
        ("作成日", created_date),
        ("有効期限", valid_until),
    ]
    for label, value in info_rows:
        ws.row_dimensions[current_row].height = 22
        lc = ws.cell(row=current_row, column=2, value=label)
        lc.font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
        lc.fill = PatternFill(fill_type="solid", fgColor=SUB_HEADER_COLOR)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = make_thin_border()

        ws.merge_cells(f"C{current_row}:G{current_row}")
        vc = ws.cell(row=current_row, column=3, value=value)
        vc.font = Font(name="Meiryo UI", size=11)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = make_thin_border()
        current_row += 1

    current_row += 1  # 空行

    # ─── 明細テーブルヘッダー ──────────────
    ws.row_dimensions[current_row].height = 28
    headers = ["No.", "工程", "作業内容", "工数 (MD)", "単価 (円)", "金額 (円)", "備考"]
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=header)
        c.font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(fill_type="solid", fgColor=HEADER_COLOR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = make_thin_border()

    table_header_row = current_row
    current_row += 1

    # ─── 明細行 ──────────────────────────
    row_no = 1
    phase_groups: dict[str, list] = {}
    for item in items_sorted:
        phase_groups.setdefault(item["phase"], []).append(item)

    for phase, phase_items in phase_groups.items():
        phase_color = PHASE_COLORS.get(phase, "F5F5F5")
        unit_price = UNIT_PRICES.get(phase, 80_000)
        phase_sub_md = 0.0
        phase_sub_amount = 0

        phase_start_row = current_row

        for item in phase_items:
            md = float(item.get("man_days", 0))
            amount = int(md * unit_price)
            phase_sub_md += md
            phase_sub_amount += amount

            ws.row_dimensions[current_row].height = 22
            apply_cell(ws, current_row, 1, row_no, align="center", fill_color=phase_color)
            apply_cell(ws, current_row, 2, phase, fill_color=phase_color)
            apply_cell(ws, current_row, 3, item.get("task", ""), fill_color=phase_color, wrap=True)
            apply_cell(ws, current_row, 4, md, align="center",
                       number_format="0.0", fill_color=phase_color)
            apply_cell(ws, current_row, 5, unit_price, align="right",
                       number_format="#,##0", fill_color=phase_color)
            apply_cell(ws, current_row, 6, amount, align="right",
                       number_format="#,##0", fill_color=phase_color)
            apply_cell(ws, current_row, 7, item.get("remarks", ""), fill_color=phase_color)

            row_no += 1
            current_row += 1

        # 工程小計行
        ws.row_dimensions[current_row].height = 24
        apply_cell(ws, current_row, 1, "", fill_color="EEEEEE")
        apply_cell(ws, current_row, 2, f"【{phase} 小計】", bold=True,
                   fill_color="EEEEEE", font_color="2E4057")
        apply_cell(ws, current_row, 3, "", fill_color="EEEEEE")
        apply_cell(ws, current_row, 4, phase_sub_md, bold=True, align="center",
                   number_format="0.0", fill_color="DDDDDD")
        apply_cell(ws, current_row, 5, "", fill_color="EEEEEE")
        apply_cell(ws, current_row, 6, phase_sub_amount, bold=True, align="right",
                   number_format="#,##0", fill_color="DDDDDD")
        apply_cell(ws, current_row, 7, "", fill_color="EEEEEE")
        current_row += 1

    current_row += 1  # 空行

    # ─── 合計行 ──────────────────────────
    total_md = sum(float(i.get("man_days", 0)) for i in items_sorted)
    total_amount = sum(
        int(float(i.get("man_days", 0)) * UNIT_PRICES.get(i["phase"], 80_000))
        for i in items_sorted
    )

    ws.row_dimensions[current_row].height = 28
    ws.merge_cells(f"A{current_row}:C{current_row}")
    tc = ws.cell(row=current_row, column=1, value="合計")
    tc.font = Font(name="Meiryo UI", size=12, bold=True, color="FFFFFF")
    tc.fill = PatternFill(fill_type="solid", fgColor=HEADER_COLOR)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = make_thin_border()
    for c in [2, 3]:
        ws.cell(row=current_row, column=c).border = make_thin_border()

    apply_cell(ws, current_row, 4, total_md, bold=True, align="center",
               number_format="0.0", fill_color=HEADER_COLOR, font_color="FFFFFF",
               font_size=12)
    apply_cell(ws, current_row, 5, "", fill_color=HEADER_COLOR)
    apply_cell(ws, current_row, 6, total_amount, bold=True, align="right",
               number_format="#,##0", fill_color=HEADER_COLOR, font_color="FFFFFF",
               font_size=12)
    apply_cell(ws, current_row, 7, "", fill_color=HEADER_COLOR)

    # ─── 注記 ────────────────────────────
    current_row += 2
    notes = [
        "※ 本見積もりは有効期限内にご発注いただいた場合に適用されます。",
        "※ 仕様変更・追加要件が発生した場合は別途ご相談ください。",
        "※ 消費税は別途申し受けます。",
    ]
    for note in notes:
        ws.merge_cells(f"A{current_row}:G{current_row}")
        nc = ws.cell(row=current_row, column=1, value=note)
        nc.font = Font(name="Meiryo UI", size=9, color="777777")
        nc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # ─── 印刷設定 ────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{table_header_row}:{table_header_row}"

    # ─── 保存 ────────────────────────────
    output_path = Path(output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "output_path": str(output_path),
        "total_man_days": total_md,
        "total_amount": total_amount,
        "item_count": len(items_sorted),
    }


# ─── エントリポイント ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="見積もり Excel ファイルを生成します")
    parser.add_argument("--input", required=True, help="入力 JSON ファイルパス")
    parser.add_argument("--output", required=True, help="出力 Excel ファイルパス (.xlsx)")
    args = parser.parse_args()

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    result = generate_excel(data, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
